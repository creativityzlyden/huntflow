import os
import argparse
import json
import sys
import openpyxl
import requests

# Взаимодействие с командной строкой
parser = argparse.ArgumentParser(description='Получение токена и пути к файлу')
parser.add_argument('token', type=str, help='Токен для доступа к hunt_flow_API')
parser.add_argument('path', type=str, help='Путь к файлу')

args = parser.parse_args()
print(args.path)

# Авторизуемся в API
url = 'https://dev-100-api.huntflow.ru'
path_to_file = os.path.join(args.path)
token = args.token

headers = {
    'Authorization': f'Bearer {token}',
    'User-Agent': 'App/1.0 (incaseoffire@example.com)',
}


# Работа с файлом
def load_data_from_file(path_to_file):
    workbook = openpyxl.load_workbook(path_to_file)
    sheet = workbook['Лист1']
    rows = sheet.max_row
    cols = sheet.max_column
    pretenders_list = []
    for row in range(2, rows + 1):
        pretenders_fields = {}
        for col in range(1, cols + 1):
            head_cell = sheet.cell(row=1, column=col)
            cell = sheet.cell(row=row, column=col)
            pretenders_fields[head_cell.value] = str(cell.value).strip()
        pretenders_list.append(pretenders_fields)
    return pretenders_list


def get_account_id():
    try:
        resp = requests.get(
            url=url + '/accounts',
            headers=headers
        ).json()
        return resp['items'][0]['id']
    except:
        print('Невозможно получить account_id')
        sys.exit(1)


def add_pretender_to_db(pretender):
    name_list = pretender['ФИО'].split()
    if len(name_list) == 3:
        patronymic = name_list[2].strip()
    else:
        patronymic = ''

    request_dict = {
        'last_name': name_list[0].strip(),
        'first_name': name_list[1].strip(),
        'middle_name': patronymic,
        'position': pretender['Должность'],
        'money': pretender['Ожидания по ЗП'],
    }

    if pretender['ИД_Фото']:
        request_dict['photo'] = pretender['ИД_Фото']

    if pretender['ИД_Файла']:
        request_dict['externals'] = [
            {
                'files': [
                    {
                        'id': pretender['ИД_Файла']
                    },
                ],
            }
        ]
    request_dict = json.dumps(request_dict)
    account_id = get_account_id()
    resp = requests.post(
        url=url + f'/account/{account_id}/applicants',
        headers=headers,
        data=request_dict
    ).json()
    pretender['ИД_Резюме'] = resp['id']


def set_status_for_pretender(pretender):
    request_dict = {
        'vacancy': pretender['Должность'],
        'status': pretender['Статус'],
        'comment': pretender['Комментарий'],
    }
    request_dict = json.dumps(request_dict)
    account_id = get_account_id()
    applicant_id = pretender['ИД_Резюме']
    requests.post(
        url=url + f'/account/{account_id}/applicants/{applicant_id}/vacancy',
        headers=headers,
        data=request_dict
    ).json()


def get_vacancies_list():
    account_id = get_account_id()
    try:
        resp = requests.get(
            url=url + f'/account/{account_id}/vacancies',
            headers=headers,
        ).json()
        return resp['items']
    except:
        print('Ошибка при получении вакансий')
        sys.exit(1)


def get_statuses_list():
    account_id = get_account_id()
    try:
        resp = requests.get(
            url=url + f'/account/{account_id}/vacancy/statuses',
            headers=headers,
        ).json()
        return resp['items']
    except:
        print('Ошибка при получении списка')
        sys.exit(1)


def add_vacancy_id_to_pretender(vacancies, pretender):
    for vacancy in vacancies:
        if pretender['Должность'] == vacancy['position']:
            pretender['ИД_Вакансии'] = vacancy['id']
            break


def add_status_id_to_pretender(statuses, pretender):
    for status in statuses:
        if pretender['Статус'] == status['name']:
            pretender['ИД_Статуса'] = status['id']
            break


def add_resume_path_to_pretender(pretender):
    folder = os.path.join(os.path.dirname(path_to_file), pretender['Должность'])
    for element in os.scandir(folder):
        if element.is_file():
            if pretender['ФИО'] in element.name:
                pretender['Путь_к_резюме'] = os.path.join(folder, element.name)
            pretender['Путь_к_резюме'] = None


def add_resume_to_db(pretender):
    file_id = None
    photo_id = None
    if pretender['Путь_к_резюме'] is None:
        pretender['ИД_Файла'] = file_id
        pretender['ИД_Фото'] = photo_id
        return
    headers_local = headers.copy()
    headers_local['X-File-Parse'] = 'true'
    file = {'file': open(pretender['Путь_к_резюме'], 'rb')}
    account_id = get_account_id()
    try:
        resp = requests.post(
            url=url + f'/account/{account_id}/upload',
            headers=headers_local,
            files=file,
        )
        if resp.status_code == 200:
            resp: resp.json()
            file_id = resp['id']
            photo_id = resp['photo']['id'] if resp['photo'] else None
        pretender['ИД_Файла'] = file_id
        pretender['ИД_Фото'] = photo_id
    except:
        print('Ошибка при загрузке файла')
        sys.exit(1)


print(f'Получение справочников')
vacancies = get_vacancies_list()
statuses = get_statuses_list()

print('Загрузка данных из файла')
pretender_list = load_data_from_file(path_to_file)


print(f'Загрузка кандидатов')
for idx, pretender in enumerate(pretender_list):
    try:
        add_vacancy_id_to_pretender(vacancies, pretender)
        add_status_id_to_pretender(statuses, pretender)
        add_resume_path_to_pretender(pretender)
        add_resume_to_db(pretender)
        add_pretender_to_db(pretender)
        set_status_for_pretender(pretender)
        print(f' - "{pretender["ФИО"]}" успешно загружен')
    except:
        load_file = 'load_resume.txt'
        print(f'Ошибка при загрузке кандидатов')
        with open(load_file, 'w', encoding='utf8') as f:
            f.write(str(idx))

print(f'Загрузка выполнена успешно')
